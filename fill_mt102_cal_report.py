# Fill MT102 Cal Report.xlsx from 2542_001_TechReview.xlsx (grid OCR of 2542_001.pdf).
import datetime
import re
import shutil

import openpyxl

SRC_GRID = "2542_001_TechReview.xlsx"
TMPL = "MT102 Cal Report.xlsx"
OUT_FILLED = "MT102 Cal Report_filled.xlsx"
TMPL_OUT = "MT102 Cal Report.xlsx"

CHISQR = {
    81: (2.613, 0.0, 1.387, 4.0),
    82: (2.071, 0.0, 1.929, 4.0),
    83: (3.716, 0.0, 0.284, 4.0),
}


def to_float(v):
    if v is None:
        return None
    t = str(v).strip()
    if not t:
        return None
    t = t.replace(",", "")
    t = re.sub(r"[\(\)\[\]lLt]+$", "", t)
    t = re.sub(r"Page\s*2.*$", "", t, flags=re.I).strip()
    m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", t)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def pass_mark(v):
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s:
        return None
    if "x" in s or "pass" in s or "\u2713" in s or "\u2714" in s:
        return "X"
    return None


def split_hi(s):
    if s is None:
        return None, None
    t = str(s).strip()
    if not t:
        return None, None
    nums = [to_float(x) for x in re.findall(r"[-+]?\d*\.?\d+", t)]
    if len(nums) >= 2:
        return nums[0], nums[1]
    if len(nums) == 1:
        return nums[0], None
    return None, None


def read_page1_body():
    wb = openpyxl.load_workbook(SRC_GRID, data_only=True)
    ws = wb["Page1"]
    start = 8
    end = ws.max_row + 1
    for r in range(start, ws.max_row + 1):
        d = ws.cell(r, 2).value
        if d and "calculat" in str(d).lower():
            end = r
            break
    rows = []
    for r in range(start, end):
        rows.append([ws.cell(r, c).value for c in range(1, 10)])
    wb.close()
    return rows


def read_page2_rows():
    wb = openpyxl.load_workbook(SRC_GRID, data_only=True)
    ws = wb["Page2"]
    rows = []
    for r in range(3, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, 12)])
    wb.close()
    return rows


def write_bk(ws, r, b, c, d, e, f, g, h, i, j, k):
    # openpyxl: ws.cell(r,c,None) does not clear a pre-existing numeric; set .value explicitly.
    for col, val in ((2, b), (3, c), (4, d), (5, e), (6, f), (7, g), (8, h), (9, i), (10, j), (11, k)):
        ws.cell(row=r, column=col).value = val


def fill_page1(ws, body):
    for idx in range(48):
        r = 5 + idx
        if idx < len(body):
            s = list(body[idx])
            while len(s) < 10:
                s.append(None)
            s = s[:10]
            h, ii = split_hi(s[7])
            if h is None:
                h = to_float(s[7])
            j = pass_mark(s[8])
            k = s[9]
            write_bk(
                ws,
                r,
                s[1],
                s[2],
                to_float(s[3]),
                to_float(s[4]),
                to_float(s[5]),
                to_float(s[6]),
                h,
                ii,
                j,
                k,
            )
        else:
            write_bk(ws, r, None, None, None, None, None, None, None, None, None, None)


def fill_test_offsets(ws, p2):
    for i in range(3):
        r = 56 + i
        if i >= len(p2):
            write_bk(ws, r, ws.cell(r, 2).value, None, None, None, None, None, None, None, None, None)
            continue
        s = p2[i]
        write_bk(
            ws,
            r,
            s[1],
            s[2],
            to_float(s[3]),
            to_float(s[4]),
            to_float(s[5]),
            to_float(s[6]),
            to_float(s[7]),
            to_float(s[8]),
            pass_mark(s[9]),
            s[10],
        )


def slice_s_matrix(p2):
    start = None
    for i, s in enumerate(p2):
        st = str(s[0]).strip() if s[0] is not None else ""
        b = str(s[1] or "").strip().lower()
        if st == "55" or b == "sxx":
            start = i
            break
    if start is None:
        return []
    out = []
    for s in p2[start:]:
        b = str(s[1] or "").strip().lower()
        if b and "cardinal" in b:
            break
        out.append(s)
        if len(out) >= 9:
            break
    return out


def slice_cardinal(p2):
    """Three rows after 'Calculating cardinal' line (xsd, ysd, zsd numbers)."""
    rows = p2
    got = []
    started = False
    for s in rows:
        b = s[1]
        bs = str(b).strip().lower() if b else ""
        if "cardinal" in bs:
            started = True
            continue
        if not started:
            continue
        if bs.startswith("e") and len(bs) <= 4:
            break
        if to_float(s[4]) is not None or to_float(s[6]) is not None:
            got.append(s)
        if len(got) >= 3:
            break
    return got[:3]


def slice_e_block(p2):
    rows = p2
    out = []
    started = False
    for s in rows:
        b = s[1]
        bs = str(b).strip().lower() if b else ""
        if bs == "exx":
            started = True
        if not started:
            continue
        if "chisqr" in bs:
            break
        if bs in ("exx", "exy", "exz", "eyx", "eyy", "eyz", "ezx", "ezy", "ezz"):
            out.append(s)
    return out[:9]


def main():
    p1 = read_page1_body()
    p2 = read_page2_rows()

    wb = openpyxl.load_workbook(TMPL)
    ws = wb["Sheet1"]

    ws["A2"] = 1475
    ws["B2"] = "2.0.0.0"
    ws["C2"] = 0
    ws["D2"] = 1
    ws["E2"] = datetime.datetime(2016, 3, 3)
    ws["F2"] = datetime.time(10, 43, 38)

    fill_page1(ws, p1)
    fill_test_offsets(ws, p2)

    smat = slice_s_matrix(p2)
    labels_s = ["sxx", "sxy", "sxz", "syx", "syy", "syz", "szx", "szy", "szz"]
    for i, lab in enumerate(labels_s):
        r = 59 + i
        ws.cell(row=r, column=2).value = lab
        if i < len(smat):
            s = smat[i]
            write_bk(
                ws,
                r,
                lab,
                s[2],
                to_float(s[3]),
                to_float(s[4]),
                to_float(s[5]),
                to_float(s[6]),
                to_float(s[7]),
                to_float(s[8]),
                pass_mark(s[9]),
                s[10],
            )
        else:
            write_bk(ws, r, lab, None, None, None, None, None, None, None, None, None)

    card = slice_cardinal(p2)
    labels_c = ["xsd", "ysd", "zsd"]
    for i, lab in enumerate(labels_c):
        r = 69 + i
        ws.cell(row=r, column=2).value = lab
        if i < len(card):
            s = card[i]
            write_bk(
                ws,
                r,
                lab,
                s[2],
                to_float(s[3]),
                to_float(s[4]),
                to_float(s[5]),
                to_float(s[6]),
                to_float(s[7]),
                to_float(s[8]),
                pass_mark(s[9]),
                s[10],
            )
        else:
            write_bk(ws, r, lab, None, None, None, None, None, None, None, None, None)

    exb = slice_e_block(p2)
    labels_e = ["exx", "exy", "exz", "eyx", "eyy", "eyz", "ezx", "ezy", "ezz"]
    for i, lab in enumerate(labels_e):
        r = 72 + i
        ws.cell(row=r, column=2).value = lab
        if i < len(exb):
            s = exb[i]
            write_bk(
                ws,
                r,
                lab,
                s[2],
                to_float(s[3]),
                to_float(s[4]),
                to_float(s[5]),
                to_float(s[6]),
                to_float(s[7]),
                to_float(s[8]),
                pass_mark(s[9]),
                s[10],
            )
        else:
            write_bk(ws, r, lab, None, None, None, None, None, None, None, None, None)

    for r in (81, 82, 83):
        lab = ws.cell(r, 2).value
        e, f, g, h = CHISQR[r]
        write_bk(ws, r, lab, None, None, e, f, g, h, None, None, None)

    wb.save(OUT_FILLED)
    try:
        shutil.copyfile(OUT_FILLED, TMPL_OUT)
        print("saved", TMPL_OUT, "and", OUT_FILLED)
    except OSError as e:
        print("saved", OUT_FILLED, "(could not overwrite", TMPL_OUT, "- close Excel and copy manually:", e, ")")
    print("p1", len(p1), "p2", len(p2), "smat", len(smat), "card", len(card), "ex", len(exb))


if __name__ == "__main__":
    main()
